import re
import os

#事件一：matplotlib.pyplot制作词云

"""
1. 清洗数据，去除表情符号、网址以及其他无意义的词语等
"""
def clean(text):
    text = re.sub(r"(回复)?(//)?\s*@\S*?\s*(:| |$)", " ", text)  # 去除正文中的@和回复/转发中的用户名
    text = re.sub(r"\[\S+\]", "", text)      # 去除表情符号
    text = re.sub(r"#\S+#", "", text)      # 保留话题内容
    URL_REGEX = re.compile(
        r'(?i)\b((?:https?://|www\d{0,3}[.]|[a-z0-9.\-]+[.][a-z]{2,4}/)(?:[^\s()<>]+|\(([^\s()<>]+|(\([^\s()<>]+\)))*\))+(?:\(([^\s()<>]+|(\([^\s()<>]+\)))*\)|[^\s`!()\[\]{};:\'".,<>?«»“”‘’]))',
        re.IGNORECASE)
    text = re.sub(URL_REGEX, "", text)       # 去除网址
    text = text.replace("转发微博", "")       # 去除无意义的词语
    text = re.sub(r"\s+", " ", text) # 合并正文中过多的空格
    return text.strip()

'''2. plt词云制作'''
from wordcloud import WordCloud
from collections import Counter
import jieba
import csv
import pandas as pd
import matplotlib.pyplot as plt
from matplotlib import colors
import numpy as np


'''词云制作前期准备'''
#读取文本
def read_it(text_path):
    with open(text_path,"r",encoding='utf-8')as f:
        lines=f.read()
        lines=lines.lower()
        text=""
        for line in lines:
            text+=line.strip()
            sentences=[doc for doc in text.split(".")]
        return text


#分词和去停用词
def cut(text):
    word_list=[]
    words=jieba.lcut(text)
    with open("stop_words.txt","r",encoding="utf-8")as f:
        stopwords=[line.strip() for line in f.readlines()]
    for word in words:
        if word not in stopwords:
            if len(word)>1:
                word_list.append(word)
    return word_list

'''统计词频和制作词云'''
def make_wc(file_get):
    if len(file_get)==0:
        tkinter.messagebox.showinfo(title='Attention', message='没有输入文件！')
    else:
        file_name=file_get
        if os.path.isfile(file_name)==0:
            tkinter.messagebox.showinfo(title='Attention', message='文件不存在！')
        else:
            '''词云制作导入'''  
            if file_name[:-3]=="csv":
                '''csv转换成txt'''
                data=pd.read_csv(file_name,encoding="utf-8")
                with open(file_name[:-3]+'txt',"a+",encoding="utf-8")as f:
                    for line in data.values:
                        clean(line[1])
                        f.write(str(line[1]))
                file_name= file_name[:-3]+'txt'      

            #词频统计
            text_path=file_name
            try:
                text_content=read_it(text_path)
            except:
                print("打开文件出错！")

            text=cut(read_it(text_path))

            count=Counter(text)
            most_count=count.most_common()


            #词云制作
            color_list=['#FF0000','#FF7F00','#000000']
            colormap=colors.ListedColormap(color_list)
            wc_coloring=np.array(plt.imread('timg.jpg'))
            wc = WordCloud(background_color = "white",\
                           font_path='C:/Windows/Fonts/simkai.ttf',\
                           #font_path="/Library/Fonts/Songti.ttc",#Mac系统
                           max_font_size=250,random_state=30,max_words=200,\
                           collocations=False,colormap=colormap,mask=wc_coloring)
            wc.generate_from_frequencies(count)
            wc.to_file(file_name[:-3]+"png")

            plt.imshow(wc)  #展示

            plt.axis("off") #横纵坐标是否显示在图上，一般要关闭

            plt.show()


#事件二：pyecharts可视化制作——微博评论情感走向图、词云、评论词频TOP10统计图

"""微博评论情感走向图前期准备——excel数据整理"""
#整理情感得分表格，转为{日期：[每条评论情感得分]}字典
class Dataclean1():            
    
    def __init__(self,filepath):
        self.filepath=filepath
    
    """对列表日期进行排序分组"""   
    def dategroup(self):
        """整理excel，只留时间和得分，写入sheet2"""
        from openpyxl import load_workbook  
        # 复制工作簿
        wb=load_workbook(self.filepath)
        ws1=wb['Sheet1']
        ws2=wb.create_sheet(title="Sheet2")
        for i,row in enumerate(ws1.iter_rows()):  
            for j,cell in enumerate(row):
                ws2.cell(row=i+1,column=j+1,value=cell.value)
        # 删除地点、评论
        ws2.delete_cols(1)
        ws2.delete_cols(2)
        wb.save(self.filepath) 
        """将工作簿2转为字典"""
        #将时间分组，每个日期对应一个情感得分列表
        import pandas as pd
        result2 = dict()
        df=pd.DataFrame(pd.read_excel(self.filepath,sheet_name="Sheet2"))
        result_dic = df.groupby('发布时间')['得分'].apply(list).to_dict()   #这里复制代码的时候要调整convert后文件中情感得分栏的标题
        #print(result_dic)
        return result_dic

#整理{日期：[情感得分]}字典，得出绘图所需数据列表
class Dataclean2():       
    def __init__(self,dic):
        self.dic=dic
    
    #整理出新列表，对应内容为：[日期]；[微博数]；[日平均得分]
    import time
    from datetime import datetime

    """日期列表"""
    def Get_date(self):
        list1=[]
        for key in self.dic.keys():
            time = key[0:10]
            list1.append(time)
        return list1
    
    """微博数列表"""
    def Get_times(self):
        list1=[]
        for value_list in self.dic.values():
            times = len(value_list)
            list1.append(times)
        return list1
    
    """日平均得分列表"""
    def Get_grade(self): 
        list1 = []
        for value_list in self.dic.values():
            grade = round(sum(value_list)/len(value_list),2)
            list1.append(grade)
        return list1

        
"""评论词频TOP10前期准备——字典数据整理"""
#将词频Top10转为作图用列表数据
class Change():
    def __init__(self,dic):
        self.dic=dic        #该字典由词云制作阶段词频统计得出
    
    """TOP10词语列表"""    
    def Get_word(self):
        list1=[]
        for key in self.dic.keys():
            word = key
            list1.append(word)
        return list1
        
    """TOP10词语对应词频"""
    def Get_value(self):
        list1=[]
        for value in self.dic.values():
            value1 = value
            list1.append(value1)
        return list1


"""pyecharts可视化制作——命令执行"""
def draw_picture(file_get):
    
    #测试输入文件名是否存在
    if len(file_get)==0:
        tkinter.messagebox.showinfo(title='Attention', message='没有输入文件！')
    else:
        file_name_before=file_get
        key=file_name_before[:-4]
        file_name=key+".xlsx"
        if os.path.isfile(file_name_before)==0:
            tkinter.messagebox.showinfo(title='Attention', message='文件不存在！')
        else:           
            from pyecharts import Line,Bar,Overlap,WordCloud,Page                    
            #from pyecharts.charts import Line,Overlap,WordCloud,Page
            # 注：需要下载pyecharts 0.5.5版本
            # pip install pyecharts==0.5.5
            
            page=Page()
            
            """微博评论情感走向图制作——折线柱状图"""
            data_before=Dataclean1(file_name).dategroup()
            data_date=Dataclean2(data_before).Get_date()
            data_times=Dataclean2(data_before).Get_times()
            data_grade=Dataclean2(data_before).Get_grade()
            #print(data_date,data_times,data_grade)
            
            bar1 = Bar("'{}'微博评论情感走向".format(key))
            bar1.add("日微博评论数",data_date,data_times,width=200,\
              bar_category_gap='30%',is_stack=True,mark_line=["average"],\
              is_datazoom_show=True,datazoom_type="slider",\
              mark_point=["max","min"],\
              mark_point_sumpol="arrow",mark_point_symbolsize=50)

            line2 = Line()
            line2.add("情感日均分",data_date,data_grade,width=200,is_stack=True,\
              is_smooth=True, mark_line=["average"],\
              is_datazoom_show=True,datazoom_type="slider",\
              mark_point=["max","min"],mark_point_sumpol="arrow",\
              mark_point_symbolsize=50)
    
            overlap = Overlap()
            overlap.add(bar1,is_add_yaxis = True)
            overlap.add(line2, yaxis_index = 1, is_add_yaxis = True)
    
            """词云制作"""
            file_name=file_get
            if os.path.isfile(file_name)==0:
                tkinter.messagebox.showinfo(title='Attention', message='文件不存在！')
            else:
                if file_name[:-3]=="csv":
                    data=pd.read_csv(file_name,encoding="utf-8",keep_default_na=False)
                    with open(file_name[:-3]+'txt',"a+",encoding="utf-8") as f:
                        for line in data.values:
                            clean(line[1])
                            f.write(str(line[1]))
                    file_name=filename[:-3]+'txt'

            #词频统计
            text_path=file_name
            try:
                text_content=read_it(text_path)
            except:
                print("打开文件出错！")

            text=cut(read_it(text_path))

            count=Counter(text)
            most_count=count.most_common()

            word=count.keys()
            value=count.values()
            #print(word,value)

            wordCloud=WordCloud()
            wordCloud.add("", word,value, word_size_range=[20, 100],shape="circle")
            
          
            """TOP10词频直方图"""
            #词频TOP10字典
            most_common_10=dict(count.most_common(10))
            #print(most_common_10)
            
            #词频TOP10列表
            word_top10=sorted(Change(most_common_10).Get_word())
            value_top10=sorted(Change(most_common_10).Get_value())
            #print(word_top10,value_top10)
            
            #统计图制作
            bar2 = Bar("评论词频TOP10")
            bar2.add("词频数",word_top10,value_top10,is_convert=True)


    """同页面集中展示三份图表：微博评论情感走向图、词云、评论词频TOP10统计图"""
    page.add(overlap)
    page.add(wordCloud)
    page.add(bar2)

    #保存html文件
    page.render("{}.html".format(key))
    print("请浏览本程序所在文件夹，打开“{}.html”文件以查看热度分析图！".format(key))
    
    








    






