#coding=utf-8

import tkinter as tk  # 使用Tkinter前需要先导入
import tkinter.messagebox # 要使用messagebox先要导入模块
from tkinter import ttk # 导入ttk模块，因为下拉菜单控件在ttk中
                        #下拉选单
import os
import csv

#运行函数
import pachong as pa
from head import headers
def spider():
    key_word=e_2.get()
    if len(e_2.get())!=0:
        start_date=cmb_y.get()+cmb_m.get().rjust(2,'0') +cmb_d.get().rjust(2,'0')
        end_date=cmb_y1.get()+cmb_m1.get().rjust(2,'0')+cmb_d1.get().rjust(2,'0')
        wbid=e_7.get()
        wbpass=e_8.get()
        print(start_date,end_date,key_word)
        
        c=choice.get()
        print(c)
        if c==2:
            pass              
            a=pa.Pacn(key_word,headers,start_date,end_date,\
                      str(key_word)+'.csv')           
            a.run()
        elif c==1:
            a=pa.Pacom(key_word,headers,start_date,end_date,str(str(key_word)+'.csv'),wbid,wbpass)           
            a.run()
        
        output_filename.set(str(key_word)+'.csv')


    else:
        tkinter.messagebox.showinfo(title='Attention', message='请输入关键词！')

import MMM

def analyze_all():
    if len(e_4.get())==0:
        tkinter.messagebox.showinfo(title='Attention', message='没有输入文件！')
    else:
        file_name=e_4.get()
        if os.path.isfile(file_name)==0:
            tkinter.messagebox.showinfo(title='Attention', message='文件不存在！')
        else:
            '''情感分析库导入'''
            if file_name[-3:]=="csv":
                '''csv转换成xlsx'''
                print("converting")
                import pandas as pd
                csv = pd.read_csv(file_name,encoding='utf-8',keep_default_na=False)
                csv.to_excel(file_name[:-3]+'xlsx', sheet_name='Sheet1')
                file_name=file_name[:-3]+'xlsx'

            '''情感分析'''
            from openpyxl import load_workbook
            from snownlp import SnowNLP
            wb=load_workbook(file_name)
            ws=wb['Sheet1']
            ws['A1']="序号"
            ws['B1']="发布时间"
            ws['C1']="发布内容"
            ws['D1']="得分"

            for cell in ws['C']:
                if cell.row==1: continue
                comment=cell.value
                MMM.clean(comment)                
                #commnet = unicode(comment, "utf-8")
                grade = SnowNLP(comment)
                print(comment, ':::',grade.sentiments) 
                ws['D'+str(cell.row)]=grade.sentiments

            wb.save(file_name)
            wb.close()   
            '''情感分析结束'''
            #MMM.make_wc(e_4.get())#词云
            MMM.draw_picture(e_4.get())#热度分析

def analyze_word():
    MMM.make_wc(e_4.get())

#界面创建
window=tk.Tk()
window.title("微博爬虫")
window.geometry('500x420')


canvas = tk.Canvas(window, bg='white',height=100, width=500)
image_file = tk.PhotoImage(file='wel.png')
image = canvas.create_image(0,0 ,anchor='nw',image=image_file)
canvas.pack()



#开始时间
f_0=tk.Frame(window)
f_0.pack()
tk.Label(f_0, text="选择开始时间", bg='white', fg='orangered',\
         font=('楷体', 15), width=20, height=1).pack(side='left') 
f_01=tk.Frame(f_0)
f_01.pack(side='right')
'''年'''
cmb_y =ttk.Combobox(f_01,width=6)
cmb_y.pack(side='left')
cmb_y['value'] = list(range(2010,2021))
cmb_y.current(0)
'''月'''
cmb_m = ttk.Combobox(f_01,width=6)
cmb_m.pack(side='left')
cmb_m['value'] = list(range(1,13))
cmb_m.current(0)
'''日'''
cmb_d = ttk.Combobox(f_01,width=6)
cmb_d.pack(side='left')
cmb_d['value'] = list(range(1,32))
cmb_d.current(0)

#结束时间
f_1=tk.Frame(window)
f_1.pack()
tk.Label(f_1, text="选择结束时间", bg='white', fg='orangered',\
         font=('楷体', 15), width=20, height=1).pack(side='left') 
f_11=tk.Frame(f_1)
f_11.pack(side='right')
'''年'''
cmb_y1 =ttk.Combobox(f_11,width=6)
cmb_y1.pack(side='left')
cmb_y1['value'] = list(range(2010,2021))
cmb_y1.current(10)
'''月'''
cmb_m1 = ttk.Combobox(f_11,width=6)
cmb_m1.pack(side='left')
cmb_m1['value'] = list(range(1,13))
cmb_m1.current(11)
'''日'''
cmb_d1 = ttk.Combobox(f_11,width=6)
cmb_d1.pack(side='left')
cmb_d1['value'] = list(range(1,32))
cmb_d1.current(30)

#选择关键词
f_2=tk.Frame(window)
f_2.pack()
tk.Label(f_2, text="输入关键词", bg='white', fg='orangered',\
         font=('楷体', 15), width=20, height=1).pack(side='left') 
f_21=tk.Frame(f_2)
f_21.pack(side='right')

e_2 = tk.Entry(f_2, width=19,font=('楷体', 14))  
e_2.pack()


#选择电脑端还是手机端
f_3=tk.Frame(window)
f_3.pack()
tk.Label(f_3, text="选择数据来源", bg='white', fg='orangered',\
         font=('楷体', 15), width=20, height=1).pack(side='left') 
f_31=tk.Frame(f_3)
f_31.pack(side='right')

choice=tk.IntVar()
radio1 = tk.Radiobutton(f_31,text="电脑端",font=('楷体', 14),value=1,variable=choice)
radio2 = tk.Radiobutton(f_31,text="手机端 ",font=('楷体', 14),value=2,variable=choice)
radio1.pack(side='left')
radio2.pack(side='right')

#输入账号
f_7=tk.Frame(window)
f_7.pack()
tk.Label(f_7, text="电脑端则输入微博账号", bg='white', fg='orangered',\
         font=('楷体', 15), width=20, height=1).pack(side='left') 
f_71=tk.Frame(f_7)
f_71.pack(side='right')

e_7 = tk.Entry(f_7, font=('楷体', 14),width=19)  
e_7.pack()

#输入密码
f_8=tk.Frame(window)
f_8.pack()
tk.Label(f_8, text="输入密码", bg='white', fg='orangered',\
         font=('楷体', 15), width=20, height=1).pack(side='left') 
f_81=tk.Frame(f_8)
f_81.pack(side='right')

e_8 = tk.Entry(f_8, font=('楷体', 14),show='*',width=19)  
e_8.pack()


#爬虫按钮
f_6=tk.Frame(window)
f_6.pack()
tk.Label(f_6, text="点击开始爬虫分析", bg='white',\
         fg='orangered', font=('楷体', 15), width=20, height=2).pack(side='left')
br_61 = tk.Button(f_6, text='开始爬虫', font=('楷体', 14),\
                width=19, height=1,command=spider)
br_61.pack(side='right')

#输入分析的文件
f_4=tk.Frame(window)
f_4.pack()
output_filename=tk.StringVar()
tk.Label(f_4, text="操作文件名", bg='white', fg='orangered',\
         font=('楷体', 15), width=20, height=1).pack(side='left') 
f_41=tk.Frame(f_4)
f_41.pack(side='right')

e_4 = tk.Entry(f_4,width=19, font=('楷体', 14),textvariable=output_filename)  
e_4.pack()

#热度分析按钮
f_9=tk.Frame(window)
f_9.pack()
tk.Label(f_9, text="生成词云", bg='white',\
         fg='orangered', font=('楷体', 15), width=20, height=2).pack(side='left')
br_9 = tk.Button(f_9, text='生成词云', font=('楷体', 14),\
                width=19, height=1,command=analyze_word)
br_9.pack(side='right')

#热度分析按钮
f_5=tk.Frame(window)
f_5.pack()
tk.Label(f_5, text="点击开始热度分析", bg='white',\
         fg='orangered', font=('楷体', 15), width=20, height=2).pack(side='left')
br_51 = tk.Button(f_5, text='开始热度分析', font=('楷体', 14),\
                width=19, height=1,command=analyze_all)
br_51.pack(side='right')


window.mainloop()
